from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook, Workbook
from tkinter import *
from webdriver_manager.chrome import ChromeDriverManager
import os
import logging

# Configuração do log
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class Aplicacao:
    def __init__(self):
        # Atributos de arquivo
        self.arquivo = 'Dados_clima.xlsx'
        self.planilha_nome = 'Lista'

        # Inicializa driver como None
        self.driver = None

        # Verifica se o arquivo Excel existe
        if not os.path.exists(self.arquivo):
            logging.info(f"Arquivo {self.arquivo} não encontrado. Criando...")
            self.criar_arquivo()

        # Inicializa os atributos de dados
        self.dt = None
        self.temp = None
        self.umi = None

        # Criação da interface gráfica
        self.criar_interface()

    def criar_interface(self):
        """Cria a interface gráfica."""
        self.layout = Tk()
        self.layout.title("Captador de Temperatura")
        self.layout.geometry("350x60")

        self.tela = Frame(self.layout)
        self.descricao = Label(self.tela, text="Clique para gerar arquivos em formato .xlsx")
        self.exportar = Button(self.tela, text="Clique aqui", command=self.capturar)

        self.tela.pack()
        self.descricao.pack()
        self.exportar.pack()

    def inicializar_driver(self):
        """Inicializa o driver do Chrome."""
        try:
            logging.info("Inicializando o ChromeDriver...")
            service = Service(ChromeDriverManager().install())
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--disable-web-security')
            options.add_argument('--disable-extensions')
            options.add_argument('--start-maximized')
            options.add_argument('--disable-infobars')
            options.add_argument(
                "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            )

            self.driver = webdriver.Chrome(service=service, options=options)
            logging.info("ChromeDriver inicializado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao inicializar o ChromeDriver: {e}")
            raise

    def criar_arquivo(self):
        """Cria um novo arquivo Excel."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self.planilha_nome
            ws.append(["Data/Hora", "Temperatura", "Umidade"])  # Cabeçalhos
            wb.save(self.arquivo)
            logging.info(f"Arquivo {self.arquivo} criado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao criar o arquivo Excel: {e}")

    def importar(self):
        """Captura informações de clima no Google."""
        try:
            if not self.driver:
                self.inicializar_driver()

            logging.info("Acessando o site para capturar dados...")
            self.driver.get('https://www.google.com.br')

            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//*[@name="q"]'))
            )

            search_box = self.driver.find_element(By.XPATH, '//*[@name="q"]')
            search_box.send_keys('clima\n')

            self.dt = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="wob_dts"]'))
            ).text
            self.temp = self.driver.find_element(By.XPATH, '//*[@id="wob_tm"]').text
            self.umi = self.driver.find_element(By.XPATH, '//*[@id="wob_hm"]').text

            logging.info(f"Dados capturados: Data/Hora: {self.dt}, Temperatura: {self.temp}°C, Umidade: {self.umi}")
        except Exception as e:
            logging.error(f"Erro ao capturar dados do site: {e}", exc_info=True)
        finally:
            if self.driver:
                self.driver.quit()
                self.driver = None
                logging.info("Navegador fechado.")

    def executar(self):
        """Atualiza a planilha com os dados capturados."""
        try:
            logging.info("Atualizando planilha...")
            tbl = load_workbook(self.arquivo)
            lista_arq = tbl[self.planilha_nome]

            proxima_linha = lista_arq.max_row + 1
            lista_arq.cell(row=proxima_linha, column=1, value=self.dt)  # Data e Hora
            lista_arq.cell(row=proxima_linha, column=2, value=self.temp)  # Temperatura
            lista_arq.cell(row=proxima_linha, column=3, value=self.umi)  # Umidade

            tbl.save(self.arquivo)
            logging.info("Planilha atualizada com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao atualizar a planilha: {e}")

    def capturar(self):
        """Executa o processo de captura e atualização."""
        self.importar()
        self.executar()

    def __del__(self):
        """Garante o encerramento do driver ao finalizar."""
        if self.driver:
            try:
                logging.info("Encerrando o ChromeDriver...")
                self.driver.quit()
                logging.info("Driver encerrado.")
            except Exception as e:
                logging.error(f"Erro ao encerrar o driver: {e}")
        else:
            logging.info("Driver não estava inicializado. Nada a encerrar.")


# Iniciar a aplicação
if __name__ == "__main__":
    try:
        app = Aplicacao()
        app.layout.mainloop()
    except Exception as e:
        logging.error(f"Erro fatal: {e}")

        
